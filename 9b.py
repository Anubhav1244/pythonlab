#create excel sheet
import pandas as pd
from openpyxl import Workbook

wb=Workbook()
sheet=wb.active
sheet.title="Language"

state=['Bihar','Jharkhand','uttarpardesh']
lang=['Hindi','Hindi','Hindi']
code=['BH','JH','UP']

sheet.cell(row=1,column=1).value="state"
sheet.cell(row=1,column=2).value="lang"
sheet.cell(row=1,column=3).value="code"

for i in range(2,5):
    sheet.cell(row=i,column=1).value=state[i-2]
    sheet.cell(row=i,column=2).value=lang[i-2]
    sheet.cell(row=i,column=3).value=code[i-2]
wb.save('wb.xlsx')
sheet=wb["Language"]

search=input("enter the code")
for i in range(2,5):
    if(sheet.cell(row=i,column=3).value==search):
        print(sheet.cell(row=i,column=1).value)
wb.close()

data=pd.read_excel('wb.xlsx')
print(data)
